"""
Parser do Cronograma Físico-Financeiro — SIENGE
Formato: Relatório SIENGE com múltiplos blocos de meses (25 colunas por bloco).

Retorna dict com:
  "obra_nome"       : str    — nome da obra
  "data_inicio"     : dict   — {"mes": int, "ano": int}
  "data_fim"        : dict   — {"mes": int, "ano": int}
  "total_obra"      : float  — orçamento total da obra
  "meses"           : list   — [(mes: int, ano: int), ...] meses cobertos no arquivo
  "custos_por_mes"  : list   — [float, ...] custo total por mês (mesma ordem de "meses")
  "contas"          : list   — [{"nome": str, "total": float, "valores": {MM/AAAA: float}}, ...]
  "data_upload"     : str    — data/hora do upload (ISO format)
  "erro"            : str    — só presente se houver falha
"""

import io
import re
import pandas as pd
from datetime import datetime

_MESES_PT = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4,
    'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8,
    'set': 9, 'out': 10, 'nov': 11, 'dez': 12
}


def _parse_mes_str(texto):
    """
    Converte 'mar/2026' → (3, 2026). Retorna (None, None) se não reconhecer.
    """
    if not texto or not isinstance(texto, str):
        return None, None
    m = re.match(r'([a-záàâãéêíóôõúç]{3})[/\s\-](\d{4})', texto.lower().strip())
    if m:
        mes = _MESES_PT.get(m.group(1)[:3])
        ano = int(m.group(2))
        if mes:
            return mes, ano
    return None, None


def _parse_data(texto):
    """
    Converte '01/07/2024' → {"mes": 7, "ano": 2024}. Retorna None se falhar.
    """
    if not texto:
        return None
    try:
        if isinstance(texto, str):
            parts = texto.strip().split('/')
            if len(parts) == 3:
                return {"mes": int(parts[1]), "ano": int(parts[2])}
    except Exception:
        pass
    return None


def parse_cronograma_sienge(data: bytes) -> dict:
    """
    Lê os bytes de um Excel SIENGE e retorna o dicionário estruturado.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {"erro": f"Não foi possível abrir o arquivo: {e}"}

    # ── 1. Encontrar todos os blocos de meses ─────────────────────────────
    # Um bloco de meses é uma linha que contém >= 2 colunas com "mes/ano"
    header_blocks = []  # lista de (linha_idx, {col_idx: (mes, ano)})
    for i, row in enumerate(all_rows):
        col_map = {}
        for j, cell in enumerate(row):
            mes, ano = _parse_mes_str(str(cell) if cell else "")
            if mes and ano:
                col_map[j] = (mes, ano)
        if len(col_map) >= 2:
            header_blocks.append((i, col_map))

    if not header_blocks:
        return {"erro": (
            "Não foram encontradas colunas de meses no arquivo. "
            "Verifique se o arquivo é um Cronograma Físico-Financeiro do SIENGE "
            "com colunas no formato 'mar/2026'."
        )}

    # ── 2. Extrair metadados da obra (da primeira seção) ──────────────────
    obra_nome = None
    data_inicio = None
    data_fim = None

    for i in range(min(15, len(all_rows))):
        row = all_rows[i]
        if row[0] == 'Obra' and len(row) > 3 and row[3]:
            obra_nome = str(row[3]).strip()
            # Data início: procura em várias colunas
            for col_idx in range(10, min(25, len(row))):
                if row[col_idx] and isinstance(row[col_idx], str) and '/' in str(row[col_idx]):
                    d = _parse_data(str(row[col_idx]))
                    if d:
                        data_inicio = d
                        break
        if row[0] == 'Unidade construtiva':
            for col_idx in range(10, min(25, len(row))):
                if row[col_idx] and isinstance(row[col_idx], str) and '/' in str(row[col_idx]):
                    d = _parse_data(str(row[col_idx]))
                    if d:
                        data_fim = d
                        break

    # ── 3. Extrair itens e valores de cada bloco ──────────────────────────
    # Estrutura: contas[nome] = {"total": float, "valores": {"MM/AAAA": float}}
    contas = {}
    total_obra = 0.0
    custos_por_mes_map = {}  # "MM/AAAA" → float total do mês

    for b_idx, (h_line, col_map) in enumerate(header_blocks):
        # Define intervalo de linhas deste bloco
        next_h = header_blocks[b_idx + 1][0] if b_idx + 1 < len(header_blocks) else len(all_rows)
        start = h_line + 2  # pula a linha de sub-cabeçalhos (ID, Descrição etc.)

        for i in range(start, next_h):
            row = all_rows[i]
            if not row or row[0] is None:
                continue

            # Linha de item de custo: col 0 é número inteiro
            if isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
                nome = str(row[1]).strip() if row[1] else ""
                if not nome or nome.lower() == 'nan':
                    continue
                # Remove prefixo "- " do nome
                nome = re.sub(r'^[\-\s]+', '', nome).strip()

                total_item = float(row[8]) if row[8] else 0.0
                if total_item == 0:
                    continue  # ignora itens sem valor (duplicatas do bloco Indiretos)

                if nome not in contas:
                    contas[nome] = {"total": total_item, "valores": {}}
                else:
                    # Atualiza total se for maior (evita duplicatas entre blocos)
                    if total_item > contas[nome]["total"]:
                        contas[nome]["total"] = total_item

                # Valores por mês
                for col_idx, (mes, ano) in col_map.items():
                    chave = f"{mes:02d}/{ano}"
                    v = row[col_idx]
                    val = float(v) if v and isinstance(v, (int, float)) else 0.0
                    if val > 0:
                        contas[nome]["valores"][chave] = (
                            contas[nome]["valores"].get(chave, 0.0) + val
                        )

            # Linha de total da obra
            elif row[0] == 'Total da obra':
                val_total = row[8]
                if val_total and float(val_total) > total_obra:
                    total_obra = float(val_total)
                # Soma valores mensais totais
                for col_idx, (mes, ano) in col_map.items():
                    chave = f"{mes:02d}/{ano}"
                    v = row[col_idx]
                    if v and isinstance(v, (int, float)) and float(v) > 0:
                        # Só registra uma vez por mês (evita duplicata entre blocos)
                        if chave not in custos_por_mes_map:
                            custos_por_mes_map[chave] = float(v)

    if not contas:
        return {"erro": (
            "Nenhuma conta de custo foi encontrada no arquivo. "
            "Verifique se o arquivo contém linhas com ID numérico e valores."
        )}

    # ── 4. Ordenar meses cronologicamente ────────────────────────────────
    def _sort_key(chave):
        mes, ano = int(chave[:2]), int(chave[3:])
        return ano * 12 + mes

    meses_ordenados = sorted(custos_por_mes_map.keys(), key=_sort_key)

    # Meses como lista de dicts
    meses_list = [{"mes": int(m[:2]), "ano": int(m[3:])} for m in meses_ordenados]
    custos_list = [custos_por_mes_map[m] for m in meses_ordenados]

    # Contas como lista
    contas_list = [
        {
            "nome": nome,
            "total": dados["total"],
            "valores": dados["valores"],
        }
        for nome, dados in contas.items()
    ]

    # Datas de início/fim a partir dos meses cobertos (fallback)
    if not data_inicio and meses_list:
        data_inicio = meses_list[0]
    if not data_fim and meses_list:
        data_fim = meses_list[-1]

    return {
        "obra_nome":      obra_nome or "Obra não identificada",
        "data_inicio":    data_inicio or {"mes": 1, "ano": 2024},
        "data_fim":       data_fim or {"mes": 12, "ano": 2026},
        "total_obra":     total_obra,
        "meses":          meses_list,
        "custos_por_mes": custos_list,
        "n_meses":        len(meses_list),
        "contas":         contas_list,
        "data_upload":    datetime.now().isoformat(),
    }
