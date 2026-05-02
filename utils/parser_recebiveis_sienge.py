"""
Parser do Relatório de Recebíveis — SIENGE
Arquivo: Contas_a_receber_-_recebiveis_*.xlsx

Estrutura:
  L9:  cabeçalho (Data vecto, Cliente, Documento, Título, Parc, TC,
                  Unid. princ, Valor original, ...)
  L10+: dados
  Col 0  (A): Data vencimento
  Col 10 (K): TC (tipo de contrato)
  Col 11 (L): Unidade
  Col 13 (N): Valor original

Tipos de TC:
  PM = Parcela Mensal        → fluxo mensal regular
  FI = Financiamento         → repasse bancário (data futura)
  CH = Cheque/À vista        → recebimento único
  RF = Reforço               → parcelas de reforço
  PC = Parcela Complementar  → complemento de parcelas
  PI = Parcela Inicial       → entrada/sinal
  PE = Permuta               → EXCLUIR (não gera caixa)

Retorna dict com:
  "obra_nome"        : str
  "data_exportacao"  : str   — data do export (do cabeçalho)
  "por_mes"          : dict  — {"AAAA-MM": float} soma de PM+FI+CH+RF+PC+PI
  "pm_por_mes"       : dict  — {"AAAA-MM": float} só parcelas mensais
  "fi_por_mes"       : dict  — {"AAAA-MM": float} só financiamentos
  "resumo_tipos"     : dict  — {TC: {"parcelas": int, "valor": float, "unidades": int}}
  "total_recebiveis" : float — sem PE
  "total_pm"         : float
  "total_fi"         : float
  "unidades_permuta" : list  — unidades com PE
  "n_unidades_pm"    : int
  "n_unidades_fi"    : int
  "arquivo_nome"     : str
  "data_upload"      : str
  "erro"             : str   — só se falhar
"""

import io
import re
from datetime import datetime
from collections import defaultdict


_TC_EXCLUIR = {"PE"}  # permuta — não gera caixa


def parse_recebiveis_sienge(data: bytes, arquivo_nome: str = "") -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {"erro": f"Não foi possível abrir o arquivo: {e}"}

    if not rows:
        return {"erro": "Arquivo vazio."}

    # Extrai nome da obra e data de exportação do cabeçalho
    obra_nome = ""
    data_exportacao = ""
    for row in rows[:8]:
        if row[0] == "Empresa" and row[1]:
            obra_nome = str(row[1]).strip()
        if row[0] and "01/" in str(row[0]):
            # Linha com período: "01/05/2026 a 01/..."
            data_exportacao = str(row[0])[:10]

    def _parse_data(v):
        if not v: return None
        if hasattr(v, 'year'):
            return v
        try:
            return datetime.strptime(str(v).strip(), "%d/%m/%Y")
        except Exception:
            return None

    # Processa linhas de dados
    por_mes          = defaultdict(float)
    pm_por_mes       = defaultdict(float)
    fi_por_mes       = defaultdict(float)
    _resumo_tipos    = {} # {tc: {"unidades": set(), "parcelas": 0, "valor": 0.0}}
    unidades_permuta = []
    _parcelas_lista  = []
    
    hoje = datetime.today()
    total_futuro = 0.0

    for row in rows[7:]: # L8 (index 7) onwards
        if not row[0]: continue
        
        _dt_raw  = str(row[0]).strip()
        _tc_raw  = str(row[10]).strip().upper() if len(row) > 10 and row[10] else ""
        _val_raw = str(row[13]).strip() if len(row) > 13 and row[13] else "0"
        _unid_r  = str(row[11]).strip() if len(row) > 11 and row[11] else ""

        # Ignorar linhas sem data válida (rodapé)
        if not any(c.isdigit() for c in _dt_raw):
            continue
            
        # PE (permuta): exclui do financeiro, mas registra unidade
        if _tc_raw == "PE":
            if _unid_r and _unid_r not in unidades_permuta:
                unidades_permuta.append(_unid_r)
            continue
            
        if _tc_raw in ("", "NAN", "A VENCER NO PERÍODO", "TOTAL DE CLIENTES", "VALOR MÉDIO"):
            continue

        try:
            _val_r = float(_val_raw.replace(',', '.'))
        except (ValueError, TypeError):
            continue
            
        if _val_r == 0:
            continue

        _dt_obj = _parse_data(row[0])
        if not _dt_obj:
            continue

        # Acumular resumo_tipos
        if _tc_raw not in _resumo_tipos:
            _resumo_tipos[_tc_raw] = {"unidades": set(), "parcelas": 0, "valor": 0.0}
        _resumo_tipos[_tc_raw]["parcelas"] += 1
        _resumo_tipos[_tc_raw]["valor"]    += _val_r
        if _unid_r and _unid_r != 'nan':
            _resumo_tipos[_tc_raw]["unidades"].add(_unid_r)

        # Acumular parcelas individuais
        _parcelas_lista.append({
            "tc":        _tc_raw,
            "valor":     _val_r,
            "data_venc": _dt_obj.strftime("%Y-%m-%d"),
            "unidade":   _unid_r if _unid_r != 'nan' else "",
        })

        # Agregações mensais
        chave = f"{_dt_obj.year}-{_dt_obj.month:02d}"
        por_mes[chave]   += _val_r
        if _tc_raw == "PM":
            pm_por_mes[chave] += _val_r
        elif _tc_raw == "FI":
            fi_por_mes[chave] += _val_r
            
        if (_dt_obj.year, _dt_obj.month) > (hoje.year, hoje.month):
            total_futuro += _val_r

    if not por_mes and not _resumo_tipos:
        return {"erro": (
            "Nenhum recebível encontrado. "
            "Verifique se é o relatório 'Contas a Receber — Recebíveis' do SIENGE."
        )}

    # Serializar resumo_tipos (sets -> int)
    resumo_final = {
        tc: {
            "unidades": len(d["unidades"]),
            "parcelas": d["parcelas"],
            "valor":    round(d["valor"], 2),
        }
        for tc, d in _resumo_tipos.items()
    }

    total_recebiveis = sum(por_mes.values())
    total_pm = sum(pm_por_mes.values())
    total_fi = sum(fi_por_mes.values())

    n_unidades_pm = len(_resumo_tipos.get("PM",{}).get("unidades",set()))
    n_unidades_fi = len(_resumo_tipos.get("FI",{}).get("unidades",set()))

    return {
        "obra_nome":         obra_nome,
        "data_exportacao":   data_exportacao,
        "por_mes":           dict(por_mes),
        "pm_por_mes":        dict(pm_por_mes),
        "fi_por_mes":        dict(fi_por_mes),
        "resumo_tipos":      resumo_final,
        "parcelas":          _parcelas_lista,
        "total_recebiveis":  total_recebiveis,
        "total_futuro":      total_futuro,
        "total_pm":          total_pm,
        "total_fi":          total_fi,
        "unidades_permuta":  unidades_permuta,
        "n_unidades_pm":     n_unidades_pm,
        "n_unidades_fi":     n_unidades_fi,
        "arquivo_nome":      arquivo_nome,
        "data_upload":       datetime.now().isoformat(),
    }
