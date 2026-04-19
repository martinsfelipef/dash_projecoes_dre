"""
Parser do Custo por Nível — SIENGE
Arquivo: emissao_custo_por_nivel-DD-MM-AAAA_-_NomeObra.xlsx

Retorna dict com:
  "obra_nome"        : str
  "periodo_final"    : str   — data do export (AAAA-MM-DD)
  "orcado_total"     : float — orçamento total da obra (DIRETOS + INDIRETOS)
  "medido_acum"      : float — valor medido acumulado total
  "realizado_acum"   : float — valor realizado acumulado total
  "comprometido"     : float — total comprometido
  "verba_disponivel" : float
  "saldo_ctp"        : float — Saldo de Contratos/Prestadores
  "pct_medido"       : float — medido/orcado*100
  "pct_realizado"    : float — realizado/orcado*100
  "cpi"              : float — medido/realizado (se realizado > 0)
  "eac"              : float — realizado + (orcado_restante / cpi)
  "custo_minimo"     : float — realizado + saldo_ctp
  "etapas_nivel2"    : list  — [{"codigo", "descricao", "orcado",
                                  "medido", "realizado", "comprometido",
                                  "verba_disp", "pct_desvio"}, ...]
  "tem_medicao"      : bool  — True se medido_acum > 0
  "arquivo_nome"     : str
  "data_upload"      : str   — ISO format
  "erro"             : str   — só presente se falhar
"""

import io
import re
from datetime import datetime


def _nivel_eap(codigo: str) -> int:
    """Retorna o nível da EAP baseado no formato do código."""
    if not codigo or not isinstance(codigo, str):
        return 0
    codigo = codigo.strip()
    # Cabeçalhos de unidade construtiva (ex: "1 - OBRA - CUSTOS DIRETOS")
    if ' - ' in codigo and not '.' in codigo:
        return 0
    partes = codigo.split('.')
    n = len(partes)
    if n == 1 and len(codigo) == 2:
        return 1   # "01"
    if n == 2:
        return 2   # "01.001"
    if n == 3:
        return 3   # "01.001.005"
    if n == 4:
        return 4   # "01.001.005.002"
    return 0


def parse_custo_nivel(data: bytes, arquivo_nome: str = "") -> dict:
    """
    Recebe bytes de um Excel CPL do SIENGE e retorna dict estruturado.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {"erro": f"Não foi possível abrir o arquivo: {e}"}

    if not rows:
        return {"erro": "Arquivo vazio."}

    # Extrai período do nome do arquivo (formato: ...-DD-MM-AAAA_-_...)
    periodo_final = ""
    if arquivo_nome:
        m = re.search(r'(\d{2})-(\d{2})-(\d{4})', arquivo_nome)
        if m:
            periodo_final = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

    # Nome da obra do nome do arquivo
    obra_nome = ""
    if arquivo_nome:
        m2 = re.search(r'_-_(.+?)\.xlsx', arquivo_nome, re.IGNORECASE)
        if m2:
            obra_nome = m2.group(1).replace('_', ' ').strip()

    # Índices das colunas (baseados no cabeçalho real)
    COL_CODIGO    = 0
    COL_DESC      = 1
    COL_UNID      = 2
    COL_ORCADO    = 3
    COL_MEDIDO    = 4
    COL_REALIZADO = 5
    COL_ESTOQUE   = 6
    COL_COMPROMETIDO = 10
    COL_PCT_DESVIO   = 11
    COL_VERBA_DISP   = 12
    COL_CTP          = 13

    def _f(v):
        """Converte para float, 0.0 se None/inválido."""
        try:
            return float(v) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    # Encontra a linha "Total da obra" (última linha com valor numérico grande)
    total_row = None
    for row in reversed(rows):
        cod = row[COL_CODIGO]
        if cod and isinstance(cod, (int, float)) and float(cod) > 1_000_000:
            total_row = row
            break

    if total_row is None:
        return {"erro": (
            "Não foi possível encontrar o 'Total da obra' no arquivo. "
            "Verifique se o arquivo é um Custo por Nível completo "
            "(com CUSTOS DIRETOS + INDIRETOS)."
        )}

    orcado_total     = _f(total_row[COL_ORCADO])
    medido_acum      = _f(total_row[COL_MEDIDO])
    realizado_acum   = _f(total_row[COL_REALIZADO])
    comprometido     = _f(total_row[COL_COMPROMETIDO])
    verba_disponivel = _f(total_row[COL_VERBA_DISP])
    saldo_ctp        = _f(total_row[COL_CTP])

    pct_medido    = (medido_acum / orcado_total * 100) if orcado_total > 0 else 0.0
    pct_realizado = (realizado_acum / orcado_total * 100) if orcado_total > 0 else 0.0
    cpi = (medido_acum / realizado_acum) if realizado_acum > 0 else 1.0
    orcado_restante = max(orcado_total - realizado_acum, 0)
    eac = realizado_acum + (orcado_restante / cpi) if cpi > 0 else orcado_total
    custo_minimo = realizado_acum + saldo_ctp

    # Coleta etapas nível 2 para tabela de desvios
    etapas_nivel2 = []
    for row in rows[1:]:  # pula cabeçalho
        cod = str(row[COL_CODIGO]).strip() if row[COL_CODIGO] else ""
        desc = str(row[COL_DESC]).strip() if row[COL_DESC] else ""

        # Ignora linhas de conciliação
        if not cod or "diferença entre" in desc.lower():
            continue
        # Ignora apropriações passadas
        if cod == "00.000.000.001":
            continue

        if _nivel_eap(cod) == 2:
            etapas_nivel2.append({
                "codigo":      cod,
                "descricao":   desc,
                "orcado":      _f(row[COL_ORCADO]),
                "medido":      _f(row[COL_MEDIDO]),
                "realizado":   _f(row[COL_REALIZADO]),
                "comprometido":_f(row[COL_COMPROMETIDO]),
                "verba_disp":  _f(row[COL_VERBA_DISP]),
                "pct_desvio":  _f(row[COL_PCT_DESVIO]),
            })

    return {
        "obra_nome":        obra_nome,
        "periodo_final":    periodo_final,
        "orcado_total":     orcado_total,
        "medido_acum":      medido_acum,
        "realizado_acum":   realizado_acum,
        "comprometido":     comprometido,
        "verba_disponivel": verba_disponivel,
        "saldo_ctp":        saldo_ctp,
        "pct_medido":       pct_medido,
        "pct_realizado":    pct_realizado,
        "cpi":              cpi,
        "eac":              eac,
        "custo_minimo":     custo_minimo,
        "etapas_nivel2":    etapas_nivel2,
        "tem_medicao":      medido_acum > 0,
        "arquivo_nome":     arquivo_nome,
        "data_upload":      datetime.now().isoformat(),
    }
