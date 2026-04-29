"""
Parser do Relatório de Unidades — SIENGE
Arquivo: Relatorio_de_unidades_*.xlsx

Estrutura:
  L5:  Empresa
  L6:  Empreendimento
  L8:  Cabeçalho (Estoque comercial, Tipo de imóvel, Nome da unidade,
                   Pavimento, Área Privativa, Área Total, Valor da unidade)
  L9+: Dados
  Col 0 (A): Status (Vendida, Disponível, Permuta, Vendido/Terceiros)
  Col 1 (B): Tipo de imóvel
  Col 2 (C): Nome da unidade
  Col 6 (G): Pavimento
  Col 18 (S): Valor da unidade

Tipos incluídos (comercializáveis):
  Apartamento, Sala, Sala Térrea, Studio, Cobertura, Loja, Flat
Tipos excluídos:
  Garagem, Vaga, Depósito, Box

Retorna dict com:
  "obra_nome"          : str
  "total_unidades"     : int  — aptos+salas SEM garagens
  "vendidas"           : int
  "permuta"            : int
  "disponiveis"        : int
  "vgv_vendido"        : float  — soma das unidades vendidas
  "preco_medio"        : float  — VGV vendido ÷ unidades vendidas
  "vgv_disponivel"     : float  — soma das unidades disponíveis
  "unidades_permuta"   : list   — nomes das unidades em permuta
  "unidades_disponiveis": list  — nomes das unidades disponíveis
  "por_status"         : dict   — {status: {"count", "valor", "unidades"}}
  "arquivo_nome"       : str
  "data_upload"        : str
  "erro"               : str    — só se falhar
"""

import io
from datetime import datetime
from collections import defaultdict

_TIPOS_INCLUIR = {
    "apartamento", "sala", "sala térrea", "studio",
    "cobertura", "loja", "flat", "kitnet"
}
_TIPOS_EXCLUIR = {
    "garagem", "vaga", "depósito", "box", "estacionamento"
}

_STATUS_VENDIDA   = {"vendida", "vendido/terceiros", "vendido terceiros"}
_STATUS_PERMUTA   = {"permuta"}
_STATUS_DISPONIVEL= {"disponível", "disponivel"}


def parse_unidades_sienge(data: bytes, arquivo_nome: str = "") -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {"erro": f"Não foi possível abrir o arquivo: {e}"}

    if not rows:
        return {"erro": "Arquivo vazio."}

    # Extrai nome da obra do cabeçalho
    obra_nome = ""
    for row in rows[:8]:
        if row[0] and str(row[0]).strip() == "Empreendimento" and row[4]:
            obra_nome = str(row[4]).strip()
            break

    por_status = defaultdict(lambda: {"count": 0, "valor": 0.0, "unidades": []})

    for row in rows[8:]:
        if not row[0] or not row[1]: continue

        status_raw = str(row[0]).strip()
        tipo_raw   = str(row[1]).strip()
        nome       = str(row[2]).strip() if row[2] else ""
        val_str    = str(row[18]).replace('R','').strip() if row[18] else "0"

        try:
            val = float(val_str)
        except Exception:
            val = 0.0

        tipo_lower   = tipo_raw.lower()
        status_lower = status_raw.lower()

        # Ignora tipos não comercializáveis
        if any(t in tipo_lower for t in _TIPOS_EXCLUIR):
            continue
        if not any(t in tipo_lower for t in _TIPOS_INCLUIR):
            continue

        por_status[status_raw]["count"]   += 1
        por_status[status_raw]["valor"]   += val
        if nome:
            por_status[status_raw]["unidades"].append(nome)

    if not por_status:
        return {"erro": (
            "Nenhuma unidade comercializável encontrada. "
            "Verifique se é o relatório 'Unidades por Empreendimento' do SIENGE."
        )}

    # Calcula totais
    total_unidades = sum(d["count"] for d in por_status.values())

    vendidas = sum(
        d["count"] for s, d in por_status.items()
        if s.lower() in _STATUS_VENDIDA
    )
    permuta = sum(
        d["count"] for s, d in por_status.items()
        if s.lower() in _STATUS_PERMUTA
    )
    disponiveis = sum(
        d["count"] for s, d in por_status.items()
        if s.lower() in _STATUS_DISPONIVEL
    )
    vgv_vendido = sum(
        d["valor"] for s, d in por_status.items()
        if s.lower() in _STATUS_VENDIDA
    )
    vgv_disponivel = sum(
        d["valor"] for s, d in por_status.items()
        if s.lower() in _STATUS_DISPONIVEL
    )
    preco_medio = vgv_vendido / vendidas if vendidas > 0 else 0.0

    unidades_permuta = []
    for s, d in por_status.items():
        if s.lower() in _STATUS_PERMUTA:
            unidades_permuta.extend(d["unidades"])

    unidades_disponiveis = []
    for s, d in por_status.items():
        if s.lower() in _STATUS_DISPONIVEL:
            unidades_disponiveis.extend(d["unidades"])

    # Serializa por_status (remove sets)
    por_status_final = {
        s: {"count": d["count"], "valor": d["valor"], "unidades": d["unidades"]}
        for s, d in por_status.items()
    }

    return {
        "obra_nome":           obra_nome,
        "total_unidades":      total_unidades,
        "vendidas":            vendidas,
        "permuta":             permuta,
        "disponiveis":         disponiveis,
        "vgv_vendido":         vgv_vendido,
        "preco_medio":         preco_medio,
        "vgv_disponivel":      vgv_disponivel,
        "unidades_permuta":    sorted(unidades_permuta),
        "unidades_disponiveis":sorted(unidades_disponiveis),
        "por_status":          por_status_final,
        "arquivo_nome":        arquivo_nome,
        "data_upload":         datetime.now().isoformat(),
    }
