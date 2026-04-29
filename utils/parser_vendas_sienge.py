"""
Parser do Relatório de Vendas — SIENGE
Arquivo: Vendas_por_Empreendimento_Simplificado_*.xlsx

Estrutura do arquivo:
  L1-9:  cabeçalho (empresa, empreendimento, período)
  L10:   header colunas
  L11+:  uma linha por unidade vendida
         Col 0  = unidade (ex: "MD APTO 703")
         Col 2  = data do contrato (str "DD/MM/AAAA")
         Col 14 = valor do contrato (int/float)
  Últimas linhas: Total Empreendimento, Contratos, Propostas, Cancelados, Total

Retorna dict com:
  "obra_nome"      : str
  "unidades_vendidas": int
  "vgv_vendido"    : float  — soma dos contratos (sem propostas/cancelados)
  "preco_medio"    : float
  "vendas_por_mes" : dict   — {"AAAA-MM": {"unidades": int, "vgv": float}}
  "data_ultima_venda": str  — "AAAA-MM-DD" da venda mais recente
  "arquivo_nome"   : str
  "data_upload"    : str    — ISO format
  "erro"           : str    — só presente se falhar
"""

import io
import re
from datetime import datetime


def parse_vendas_sienge(data: bytes, arquivo_nome: str = "") -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {"erro": f"Não foi possível abrir o arquivo: {e}"}

    if not rows:
        return {"erro": "Arquivo vazio."}

    # Detecta prefixo do empreendimento (MD, TC, etc.) da linha 11
    prefixo = ""
    obra_nome = ""
    for row in rows[5:9]:
        if row[0] == "Empreendimento" and row[1]:
            obra_nome = str(row[1]).strip()
            break

    # Detecta prefixo da primeira linha de dados (col 0 = "MD APTO 703")
    for row in rows[10:15]:
        if row[0] and isinstance(row[0], str) and len(row[0]) >= 2:
            prefixo = row[0][:2].upper()
            break

    def _parse_data(v):
        if not v: return None
        try:
            return datetime.strptime(str(v).strip(), '%d/%m/%Y')
        except Exception:
            return None

    vendas = []
    for row in rows[10:]:
        unidade = row[0]
        if not unidade or not isinstance(unidade, str): continue
        if prefixo and not str(unidade).upper().startswith(prefixo): continue
        data_str = row[2]
        valor    = row[14]
        if not valor or not isinstance(valor, (int, float)): continue
        dt = _parse_data(data_str)
        if dt is None: continue
        mes_ano = f"{dt.year}-{dt.month:02d}"
        vendas.append({
            "unidade":  str(unidade).strip(),
            "data":     dt.strftime("%Y-%m-%d"),
            "mes_ano":  mes_ano,
            "valor":    float(valor),
        })

    if not vendas:
        return {"erro": (
            "Nenhuma venda encontrada no arquivo. "
            "Verifique se é um relatório 'Vendas por Empreendimento — Simplificado' do SIENGE."
        )}

    # Agrupa por mês
    vendas_por_mes = {}
    for v in vendas:
        m = v["mes_ano"]
        if m not in vendas_por_mes:
            vendas_por_mes[m] = {"unidades": 0, "vgv": 0.0}
        vendas_por_mes[m]["unidades"] += 1
        vendas_por_mes[m]["vgv"]      += v["valor"]

    unidades_vendidas = len(vendas)
    vgv_vendido       = sum(v["valor"] for v in vendas)
    preco_medio       = vgv_vendido / unidades_vendidas if unidades_vendidas > 0 else 0.0
    data_ultima_venda = max(v["data"] for v in vendas) if vendas else ""

    return {
        "obra_nome":          obra_nome,
        "unidades_vendidas":  unidades_vendidas,
        "vgv_vendido":        vgv_vendido,
        "preco_medio":        preco_medio,
        "vendas_por_mes":     vendas_por_mes,
        "data_ultima_venda":  data_ultima_venda,
        "arquivo_nome":       arquivo_nome,
        "data_upload":        datetime.now().isoformat(),
    }
